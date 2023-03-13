import tkinter as tk
import os
from openpyxl import *

window=tk.Tk()

window.title("Klasör Güncelle")
window.geometry("650x450")

label1=tk.Label(window,text="Proje excel dosyalarının bulunduğu dizini giriniz")
label1.pack()

entry1=tk.Entry(window,bd=5)
entry1.pack()

label2=tk.Label(window,text="Proje kodunu giriniz")
label2.pack()

entry2=tk.Entry(window,bd=5)
entry2.pack()

list_harfler=["A","B","C","D","E","F","G","H","I","İ","J","K","L","M","N","O","P","R","S","T","U","V","Y","Z"]

def buton():
    wb1=Workbook()
    filepath=entry1.get() + "\\" + "Sipariş_Çalışma_Hazırlığı_" + entry2.get() + ".xlsx" #Proje kodu bilerek sonda.For dönerken veri aktarılacak excel dosyası almaması için. 
    wb1.save(filepath)
    wb1.close()

    wb4=load_workbook(filepath)
    sheet=wb4.active
    sheet.append(("Ürün Kodu","İndis","Malzeme Tanım ","Kalınlık","Boya","En","Boy","Parça (metrekare) ","Boşaltma (metrekare) ","Hammadde Cinsi","Büküm","Saç Plaka En","Saç Plaka Boy","Yerleşim Sayısı","İşleme Süresi","Hurda Ağırlık"))
    wb4.save(filepath)
    wb4.close()

    os.chdir(entry1.get())
    dosyalar_list=os.listdir(entry1.get())
    for i in dosyalar_list:
        dosya_split=i.split("_")
        if dosya_split[0]== entry2.get():
            wb2=load_workbook(entry1.get() + "\\" + i)
            ws=wb2.active
            sütun_urun_kodu= ws["B11"].value
            sütun_urun_kodu_new=sütun_urun_kodu[0:7]
            
            if sütun_urun_kodu[7] in list_harfler:
                sütun_indis=sütun_urun_kodu[7]
            else:
                sütun_indis="BOŞ"

            sütun_kalınlık=ws["B3"].value
            sütun_kalınlık_new=sütun_kalınlık.split(" ")
            sütun_kalınlık_new_1=sütun_kalınlık_new[1][1:]
            sütun_en=ws["F11"].value
            sütun_boy=ws["E11"].value
            sütun_hammadde_cinsi=ws["B3"].value
            sütun_sac_plaka_en=ws["B4"].value
            sütun_sac_plaka_en_new=sütun_sac_plaka_en.split(" ")
            sütun_sac_plaka_en_new_1=sütun_sac_plaka_en_new[2]
            sütun_sac_plaka_boy=sütun_sac_plaka_en_new[0]         
            sütun_yerlesim_sayısı=ws["H11"].value
            sütun_isleme_süresi=ws["B5"].value
            sütun_hurda_agırlık=ws["D7"].value
            wb2.close()

            wb3=load_workbook(filepath)
            sheet=wb3.active
            sheet.append((sütun_urun_kodu_new,sütun_indis," ",sütun_kalınlık_new_1," ",sütun_en,sütun_boy," "," ",sütun_hammadde_cinsi," ",sütun_sac_plaka_en_new_1,sütun_sac_plaka_boy,sütun_yerlesim_sayısı,sütun_isleme_süresi,sütun_hurda_agırlık))
            wb3.save(filepath)
            wb3.close()

button=tk.Button(window,text="Gönder",command=buton)
button.pack()

window.mainloop()